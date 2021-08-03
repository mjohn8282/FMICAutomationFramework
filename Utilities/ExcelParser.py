import openpyxl



def getexcelrowcount(filename,sheetname):
    workbook = openpyxl.load_workbook(filename, sheetname)
    row = workbook[sheetname]
    return row.max_row

def getexcelcolumncount(filename,sheetname):
    workbook = openpyxl.load_workbook(filename, sheetname)
    col = workbook[sheetname]
    return col.max_column

def reader(filename,sheetname,rownum,colnum):
    workbook=openpyxl.load_workbook(filename,sheetname)
    sheeet=workbook[sheetname]
    data= sheeet.cell(row=rownum,column=colnum)
    return data.value









